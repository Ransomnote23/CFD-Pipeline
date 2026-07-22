from http.server import BaseHTTPRequestHandler
import json, io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ────────────────────────────────────────────────────────────────────
def today_str():
    return date.today().strftime("%B %d, %Y").replace(" 0"," ")

def parse_date(s):
    if not s: return None
    try: return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except: return None

def fmt_date(s):
    d = parse_date(s)
    return d.strftime("%Y-%m-%d") if d else "—"

def is_overdue(due_str):
    d = parse_date(due_str)
    return d is not None and d < date.today()

def line_shipped_qty(line):
    return sum(s.get("qty",0) for s in line.get("shipped",[]))

def line_open_qty(line):
    return line["qty"] - line_shipped_qty(line)

def line_status(line):
    shipped = line_shipped_qty(line)
    open_qty = line["qty"] - shipped
    if open_qty <= 0:
        has_inv = any(s.get("inv") for s in line.get("shipped",[]))
        return "INVOICED" if has_inv else "COMPLETE"
    return "PARTIAL" if shipped > 0 else "OPEN"

# ── styles ─────────────────────────────────────────────────────────────────────
FN = "Arial"
def fill(hex): return PatternFill("solid", fgColor=hex)
def font(bold=False, color="000000", size=9, italic=False):
    return Font(name=FN, bold=bold, color=color, size=size, italic=italic)

thin = Side(style="thin",   color="CCCCCC")
med  = Side(style="medium", color="888888")
thk  = Side(style="medium", color="1F4E79")

BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_MED   = Border(left=med,  right=med,  top=med,  bottom=med)
BORDER_HEAVY = Border(left=thk,  right=thk,  top=thk,  bottom=thk)
BORDER_BOT   = Border(bottom=Side(style="medium", color="1F4E79"))

F_HDR   = fill("1F4E79")  # navy title bar
F_COL   = fill("2E75B6")  # blue col headers
F_PO    = fill("D6E4F0")  # light blue PO grouping row
F_OD    = fill("FFD7D7")  # red overdue
F_OD_PO = fill("FFBCBC")  # darker red overdue PO header
F_ALT   = fill("F0F6FF")  # alternating row
F_TOT   = fill("E2EFDA")  # green totals
F_PROJ  = fill("FFFDE7")  # yellow proj delivery col
F_WHITE = fill("FFFFFF")
F_LGND  = fill("F5F5F5")  # legend bg

al_c = Alignment(horizontal="center", vertical="center")
al_l = Alignment(horizontal="left",   vertical="center")
al_r = Alignment(horizontal="right",  vertical="center")
al_w = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def sc(ws, r, c, v, fnt=None, fil=None, aln=None, brd=None, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    if fnt: cell.font      = fnt
    if fil: cell.fill      = fil
    if aln: cell.alignment = aln
    if brd: cell.border    = brd
    if nf:  cell.number_format = nf
    return cell

# ── summary sheet ──────────────────────────────────────────────────────────────
SUM_HDRS   = ["Customer PO#","WO#","Date Received","Due Date",
               "Total Lines","Open Lines","Partial Lines","Open Value ($)"]
SUM_WIDTHS = [20, 12, 16, 16, 13, 12, 14, 16]

def build_summary(ws, customer, po_meta):
    ws.freeze_panes = "A6"

    # Row 1 — title bar
    ws.merge_cells("A1:H1")
    sc(ws,1,1, f"Open Purchase Order Status Report  ·  As of {today_str()}",
       fnt=font(bold=True,color="FFFFFF",size=14),
       fil=F_HDR, aln=Alignment(horizontal="center",vertical="center"))
    ws.row_dimensions[1].height = 28

    # Row 2 — customer name
    ws.merge_cells("A2:H2")
    sc(ws,2,1, customer,
       fnt=font(bold=True,color="1F4E79",size=14),
       aln=al_l)
    ws.row_dimensions[2].height = 22

    # Row 3 — spacer
    ws.row_dimensions[3].height = 5

    # Row 4 — col headers
    for ci,h in enumerate(SUM_HDRS,1):
        sc(ws,4,ci,h, fnt=font(bold=True,color="FFFFFF",size=9),
           fil=F_COL, aln=al_c, brd=BORDER)
    ws.row_dimensions[4].height = 18

    total_open = total_pd = 0
    r = 5
    for i, po in enumerate(po_meta):
        od  = po["od"]
        fil = F_OD if od else (F_ALT if i%2==0 else F_WHITE)
        fc  = "BB0000" if od else "000000"
        vals = [po["cpo"], po["wo"],
                parse_date(po["rcvd"]), parse_date(po["due"]),
                po["total"], po["open_l"], po["part_l"], po["open_v"]]
        for ci, v in enumerate(vals, 1):
            nf=None; aln=al_c
            if ci==1: aln=al_l
            if ci==8: nf='#,##0.00'; aln=al_r
            if ci in (3,4): nf='YYYY-MM-DD'
            sc(ws,r,ci,v, fnt=font(color=fc,size=9),
               fil=fil, aln=aln, brd=BORDER, nf=nf)
        total_open += po["open_v"]
        if od: total_pd += po["open_v"]
        r += 1

    # Totals rows
    def tot_row(row, label, val, fil, fc):
        sc(ws,row,1,label, fnt=font(bold=True,color=fc,size=9), fil=fil, aln=al_l, brd=BORDER)
        for ci in range(2,8):
            sc(ws,row,ci,"", fil=fil, brd=BORDER)
        sc(ws,row,8,val, fnt=font(bold=True,color=fc,size=9),
           fil=fil, aln=al_r, brd=BORDER, nf='#,##0.00')

    tot_row(r,   "TOTAL",          total_open, F_TOT, "1F6B00"); r+=1
    tot_row(r,   "TOTAL PAST DUE", total_pd,   F_OD,  "BB0000"); r+=2

    # Legend
    sc(ws,r,1,"Status Legend:", fnt=font(bold=True,size=8,color="555555"), fil=F_LGND, aln=al_l)
    for ci in range(2,9): sc(ws,r,ci,"", fil=F_LGND)
    r+=1
    for code, desc, fc in [
        ("OPEN",     "Items not yet shipped",  "000000"),
        ("PARTIAL",  "Partially shipped",       "CC5500"),
        ("Red row",  "Due date has passed",     "BB0000"),
    ]:
        sc(ws,r,1,code, fnt=font(bold=True,size=8,color=fc)); 
        sc(ws,r,2,desc, fnt=font(size=8,color="555555"))
        r+=1

    for ci,w in enumerate(SUM_WIDTHS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── detail sheet ───────────────────────────────────────────────────────────────
DET_HDRS = ["Customer PO#","WO#","Line","Description","Ordered Qty",
            "Shipped Qty","Open Qty","Status","Proj. Delivery Date ✎",
            "Labor Hours","Due Date","Date Received"]
DET_WIDTHS = [18,11,7,46,13,12,10,10,21,13,14,14]

STATUS_COLORS = {
    "OPEN":"CC0000","PARTIAL":"CC5500",
    "INVOICED":"1A7A1A","COMPLETE":"1A7A1A"
}

def build_detail(ws, customer, po_list, skus_map):
    ws.freeze_panes = "A6"

    # Title
    ws.merge_cells("A1:L1")
    sc(ws,1,1, f"{customer} — Open PO Line Item Detail",
       fnt=font(bold=True,color="FFFFFF",size=14),
       fil=F_HDR, aln=Alignment(horizontal="center",vertical="center"))
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:L2")
    sc(ws,2,1,
       f"Report Date: {today_str()}    ·    "
       f"Labor Hours = BOM hrs/unit × Open Qty × 1.2 safety factor    ·    "
       f"'Proj. Delivery Date' column is blank — fill in by hand",
       fnt=font(size=8,italic=True,color="555555"), aln=al_w)
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 5

    # Column headers
    for ci,h in enumerate(DET_HDRS,1):
        f_use = F_PROJ if ci==9 else F_COL
        fc_use= "666666" if ci==9 else "FFFFFF"
        sc(ws,4,ci,h, fnt=font(bold=True,color=fc_use,size=9),
           fil=f_use, aln=al_c, brd=BORDER)
    ws.row_dimensions[4].height = 30

    r = 5
    for po in po_list:
        od = is_overdue(po.get("due",""))
        po_start_row = r  # remember where this PO block starts

        # PO header row — bold navy or dark crimson, white text
        ws.merge_cells(f"A{r}:L{r}")
        po_lbl = (f"PO# {po['cpo']}    ·    WO# {po.get('wo') or '—'}    ·    "
                  f"Due: {fmt_date(po.get('due'))}    ·    "
                  f"Received: {fmt_date(po.get('rcvd'))}")
        po_hdr_fill = fill("8B0000") if od else fill("1F3864")  # dark crimson : dark navy
        sc(ws,r,1, po_lbl,
           fnt=font(bold=True, color="FFFFFF", size=10),
           fil=po_hdr_fill,
           aln=al_l, brd=BORDER_MED)
        ws.row_dimensions[r].height = 18
        r += 1

        for li, line in enumerate(po.get("lines",[])):
            shq = line_shipped_qty(line)
            oq  = line_open_qty(line)
            st  = line_status(line)
            bom = skus_map.get(line["desc"],0)
            lh  = round(bom*oq*1.2,2) if (bom and oq>0) else None

            row_od = od and oq>0
            fil = F_OD if row_od else (F_ALT if li%2==0 else F_WHITE)

            vals=[po["cpo"], po.get("wo") or "—", line["id"], line["desc"],
                  line["qty"], shq, oq, st, None, lh,
                  parse_date(po.get("due")), parse_date(po.get("rcvd"))]

            for ci, v in enumerate(vals,1):
                nf=None; aln=al_c
                fnt_use = font(color=STATUS_COLORS.get(st,"000000") if ci==8 else
                               ("BB0000" if row_od and ci!=9 else "000000"), size=9)
                if ci==4:  aln=al_l
                if ci in (5,6,7): nf="0"; aln=al_r
                if ci==10: nf="0.00"; aln=al_r
                if ci in (11,12): nf="YYYY-MM-DD"; aln=al_c
                if ci==9:  aln=al_c; fnt_use=font(color="AAAAAA",size=8,italic=True)
                fil_use = F_PROJ if ci==9 else fil
                sc(ws,r,ci,v, fnt=fnt_use, fil=fil_use, aln=aln, brd=BORDER, nf=nf)

            ws.row_dimensions[r].height = 14
            r += 1

        po_end_row = r - 1  # last line row of this PO

        # ── Bold border box around entire PO block (header + all lines) ──────
        box_color = "8B0000" if od else "1F3864"
        s_thick = Side(style="medium", color=box_color)
        s_none  = Side(style=None)
        for row_i in range(po_start_row, po_end_row + 1):
            is_top    = (row_i == po_start_row)
            is_bottom = (row_i == po_end_row)
            for col_i in range(1, 13):
                is_left  = (col_i == 1)
                is_right = (col_i == 12)
                cell = ws.cell(row=row_i, column=col_i)
                # Build border: thick on box edges, keep existing thin between cells
                top_side    = s_thick if is_top    else cell.border.top
                bottom_side = s_thick if is_bottom else cell.border.bottom
                left_side   = s_thick if is_left   else cell.border.left
                right_side  = s_thick if is_right  else cell.border.right
                cell.border = Border(top=top_side, bottom=bottom_side,
                                     left=left_side, right=right_side)

        # Spacer between POs
        ws.merge_cells(f"A{r}:L{r}")
        sc(ws,r,1,"", fil=F_WHITE)
        ws.row_dimensions[r].height = 8
        r += 1

    for ci,w in enumerate(DET_WIDTHS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── main builder ───────────────────────────────────────────────────────────────
def build_workbook(data, customer_filter=None):
    pos      = data["pos"]
    skus     = data.get("skus",[])
    skus_map = {s["desc"]: s.get("laborHours",0) for s in skus}

    from collections import defaultdict
    groups = defaultdict(list)
    for po in pos:
        if not any(line_open_qty(l)>0 for l in po.get("lines",[])):
            continue
        c = po.get("customer","Unknown")
        if customer_filter and c != customer_filter:
            continue
        groups[c].append(po)

    customers = sorted(groups.keys())
    if not customers:
        raise ValueError("No open POs found.")

    for c in customers:
        groups[c].sort(key=lambda p:(
            0 if is_overdue(p.get("due","")) else 1,
            p.get("due","9999") or "9999"
        ))

    wb = Workbook()
    wb.remove(wb.active)

    # Build metadata for all customers first
    all_meta = {}
    for cust in customers:
        po_list = groups[cust]
        po_meta = []
        for po in po_list:
            total=len(po.get("lines",[])); openl=partl=0; openv=0.0
            for l in po.get("lines",[]):
                oq=line_open_qty(l)
                if oq>0:
                    s=line_status(l)
                    if s=="PARTIAL": partl+=1
                    else: openl+=1
                    openv+=oq*l.get("unitPrice",0)
            po_meta.append(dict(
                cpo=po["cpo"], wo=po.get("wo") or "—",
                rcvd=po.get("rcvd",""), due=po.get("due",""),
                total=total, open_l=openl, part_l=partl, open_v=openv,
                od=is_overdue(po.get("due",""))
            ))
        all_meta[cust] = po_meta

    # Pass 1 — all Summary tabs
    for cust in customers:
        safe = cust[:24]
        ws_sum = wb.create_sheet(f"{safe[:27]} Sum")
        build_summary(ws_sum, cust, all_meta[cust])

    # Pass 2 — all Detail tabs
    for cust in customers:
        safe = cust[:24]
        ws_det = wb.create_sheet(f"{safe[:24]} Detail")
        build_detail(ws_det, cust, groups[cust], skus_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Vercel handler ─────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            length  = int(self.headers.get("Content-Length",0))
            body    = self.rfile.read(length)
            payload = json.loads(body)
            cust    = payload.get("customer") or None
            data    = payload["data"]

            xlsx_bytes = build_workbook(data, cust)

            self.send_response(200)
            self._cors()
            self.send_header("Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",
                f'attachment; filename="Open_PO_Report.xlsx"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            self.send_response(500)
            self._cors()
            self.send_header("Content-Type","application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error":str(e)}).encode())

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Access-Control-Allow-Methods","POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers","Content-Type")

    def log_message(self, *a): pass
